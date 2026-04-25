import requests
import openpyxl
from openpyxl.styles import Font, Alignment
import time

# Dicionário para salvar (fazer cache) dados dos Pokémon e evitar baixar a mesma coisa duas vezes
pokemon_cache = {}

def get_pokemon_data(pid):
    """Busca os dados de um Pokémon específico, focado no HGSS."""
    # Ignora Pokémon da Geração 5 em diante (ID > 493)
    if pid > 493:
        return None
    
    if pid in pokemon_cache:
        return pokemon_cache[pid]
        
    url = f"https://pokeapi.co/api/v2/pokemon/{pid}"
    try:
        res = requests.get(url)
        if res.status_code != 200:
            return None
        data = res.json()
    except:
        return None

    name = data['name'].capitalize()
    moves = []
    
    for item in data['moves']:
        for v in item['version_group_details']:
            if v['version_group']['name'] == 'heartgold-soulsilver' and v['move_learn_method']['name'] == 'level-up':
                lvl = v['level_learned_at']
                if lvl > 0: # Ignora golpes de nível 0
                    moves.append((lvl, item['move']['name'].title().replace('-', ' ')))
                    
    # Ordena os golpes pelo nível
    moves.sort(key=lambda x: x[0])
    
    pokemon_cache[pid] = {'id': pid, 'name': name, 'moves': moves}
    return pokemon_cache[pid]

def get_evolution_paths(chain_node):
    """Navega pela árvore de evolução e retorna todas as rotas possíveis (ex: Eevee -> Vaporeon)."""
    species_url = chain_node['species']['url']
    # Extrai o ID do Pokémon a partir da URL da espécie
    species_id = int(species_url.strip('/').split('/')[-1])
    
    # Se não tem evolução, a rota acaba aqui
    if not chain_node['evolves_to']:
        return [[species_id]]
        
    paths = []
    for evo in chain_node['evolves_to']:
        sub_paths = get_evolution_paths(evo)
        for p in sub_paths:
            paths.append([species_id] + p)
            
    return paths

def gerar_planilha_horizontal():
    print("Iniciando a montagem da planilha horizontal pelas cadeias evolutivas...")
    print("Isso vai levar um tempo (cerca de 5 a 10 minutos) pois a API tem limites de velocidade.\n")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Familias HGSS"
    
    # Adicionando os cabeçalhos para até 3 estágios de evolução (12 colunas)
    headers = ['ID', 'Pokemon', 'Level', 'Golpe'] * 3
    ws.append(headers)
    
    # Deixando o cabeçalho em negrito
    for col_num in range(1, 13):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    current_row = 2
    
    # Gen 4 (até Arceus) está contida nas primeiras ~250 evolution chains
    for chain_id in range(1, 260):
        url = f"https://pokeapi.co/api/v2/evolution-chain/{chain_id}/"
        try:
            res = requests.get(url)
            if res.status_code != 200:
                continue
            chain_data = res.json()
        except:
            continue
            
        # Pega todas as rotas evolutivas dessa família
        paths = get_evolution_paths(chain_data['chain'])
        
        for path_ids in paths:
            # Baixa os dados de todos os Pokémon na rota
            path_data = []
            for pid in path_ids:
                data = get_pokemon_data(pid)
                if data: # Só adiciona se for Gen 4 ou anterior
                    path_data.append(data)
            
            if not path_data:
                continue
                
            print(f"Processando família: {' -> '.join([p['name'] for p in path_data])}")
            
            # Descobre qual Pokémon dessa linha tem a maior lista de golpes para definir a altura do bloco
            max_moves_len = max([len(p['moves']) for p in path_data] + [1])
            
            # Preenche as colunas para cada estágio (Base = Col 1, Estágio 1 = Col 5, etc)
            for stage_idx, pkmn in enumerate(path_data):
                col_base = (stage_idx * 4) + 1 # 1, 5, 9...
                
                # Preenche a altura total do bloco (mesmo que os golpes já tenham acabado)
                for r_offset in range(max_moves_len):
                    row_idx = current_row + r_offset
                    
                    # Coluna ID e Nome sempre são preenchidas para alinhar visualmente
                    ws.cell(row=row_idx, column=col_base, value=pkmn['id']).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_idx, column=col_base + 1, value=pkmn['name'])
                    
                    # Coluna Level e Golpe só preenche se ainda houver golpe na lista
                    if r_offset < len(pkmn['moves']):
                        lvl, move_name = pkmn['moves'][r_offset]
                        ws.cell(row=row_idx, column=col_base + 2, value=lvl).alignment = Alignment(horizontal='right')
                        ws.cell(row=row_idx, column=col_base + 3, value=move_name)
            
            # Pula para a próxima linha livre, de acordo com o tamanho do bloco que acabou de ser desenhado
            current_row += max_moves_len
            
        # Pausa para não sobrecarregar a PokéAPI
        time.sleep(0.1)

    # Ajusta o tamanho das colunas para ficar legível
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    nome_arquivo = 'Golpes_Familias_HGSS.xlsx'
    wb.save(nome_arquivo)
    print(f"\n🎉 SUCESSO! Arquivo '{nome_arquivo}' gerado exatamente no formato solicitado!")

if __name__ == "__main__":
    gerar_planilha_horizontal()