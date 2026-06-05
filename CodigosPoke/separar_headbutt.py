import openpyxl
import re
import copy

# Nome do seu arquivo (ajuste se necessário)
ARQUIVO_EXCEL = "SoulSilverOficial.xlsm"
NOME_ABA = "WildPokemon" # <--- MUDE PARA O NOME EXATO DA SUA ABA!

print("Carregando planilha... Isso pode levar alguns segundos.")
# keep_vba=True é VITAL para não apagar suas macros!
wb = openpyxl.load_workbook(ARQUIVO_EXCEL, keep_vba=True)

# Se quiser usar a aba ativa, comente a linha abaixo e use: ws = wb.active
ws = wb[NOME_ABA]

max_row = ws.max_row
linhas_modificadas = 0

print("Analisando as linhas de Headbutt (de baixo para cima)...")

# Iteramos de trás para frente! (Assim as inserções não bagunçam o index do loop)
for row in range(max_row, 1, -1):
    celula_local = ws.cell(row=row, column=3).value
    
    # Verifica se a linha tem Headbutt
    if celula_local and "Headbutt" in str(celula_local):
        
        # Lê os dados dos slots D(4) até U(21)
        slots_data = []
        for col in range(4, 22):
            cell = ws.cell(row=row, column=col)
            # Guardamos o valor e a formatação (cores de fundo, fonte, etc)
            slots_data.append({
                'value': cell.value,
                'fill': copy.copy(cell.fill) if cell.fill else None,
                'font': copy.copy(cell.font) if cell.font else None,
                'border': copy.copy(cell.border) if cell.border else None,
                'alignment': copy.copy(cell.alignment) if cell.alignment else None
            })
            
        # Conta quantos Pokémons existem de fato nessa linha
        qtd_pokemons = sum(1 for p in slots_data if p['value'] is not None and str(p['value']).strip() != "")
        
        if qtd_pokemons > 6:
            # Define quantos grupos precisamos (1 a 6 = 1 grupo | 7 a 12 = 2 grupos | 13 a 18 = 3 grupos)
            num_grupos = 2 if qtd_pokemons <= 12 else 3
            linhas_a_inserir = num_grupos - 1
            
            # 1. Insere as linhas em branco logo abaixo da atual
            ws.insert_rows(row + 1, amount=linhas_a_inserir)
            
            # 2. DESLOCA OS IDs PARA BAIXO
            # Como inserimos linhas, todas as linhas abaixo das inseridas precisam ter o ID somado
            for r_shift in range(row + num_grupos, ws.max_row + 1):
                id_cell = ws.cell(row=r_shift, column=2)
                if id_cell.value is not None and isinstance(id_cell.value, (int, float)):
                    id_cell.value = int(id_cell.value) + linhas_a_inserir
                    
            # 3. PREENCHE OS DADOS NOS GRUPOS
            original_id = int(ws.cell(row=row, column=2).value)
            nome_base = str(celula_local)
            
            for g in range(num_grupos):
                curr_row = row + g
                
                # Arruma o ID do novo grupo
                ws.cell(row=curr_row, column=2).value = original_id + g
                
                # Arruma o Nome: Troca "Headbutt" por "Headbutt GrpX"
                # Exemplo: "Vila Primavera Headbutt (rate=1)" vira "Vila Primavera Headbutt Grp1 (rate=1)"
                novo_nome = re.sub(r'(Headbutt)', rf'\1 Grp{g+1}', nome_base)
                ws.cell(row=curr_row, column=3).value = novo_nome
                
                # Pega formatação visual das colunas A, B e C da linha original
                if g > 0:
                    for col_abc in range(1, 4):
                        fonte = ws.cell(row=row, column=col_abc)
                        alvo = ws.cell(row=curr_row, column=col_abc)
                        if col_abc == 1: alvo.value = fonte.value # Copia valor da checkbox/coluna A
                        if fonte.fill: alvo.fill = copy.copy(fonte.fill)
                        if fonte.font: alvo.font = copy.copy(fonte.font)
                        if fonte.alignment: alvo.alignment = copy.copy(fonte.alignment)
                
                # Escreve apenas 6 Pokémons nas colunas D(4) até I(9)
                inicio_idx = g * 6
                fim_idx = inicio_idx + 6
                
                col_insert = 4
                for p_idx in range(inicio_idx, fim_idx):
                    target_cell = ws.cell(row=curr_row, column=col_insert)
                    
                    if p_idx < len(slots_data):
                        target_cell.value = slots_data[p_idx]['value']
                        # Traz a cor exata (amarelo, verde, etc) daquele slot específico
                        if slots_data[p_idx]['fill']: target_cell.fill = slots_data[p_idx]['fill']
                        if slots_data[p_idx]['font']: target_cell.font = slots_data[p_idx]['font']
                    else:
                        target_cell.value = None
                        target_cell.fill = openpyxl.styles.PatternFill(fill_type=None)
                        
                    col_insert += 1
                
                # Apaga os dados das colunas J(10) até U(21) dessa linha para limpar a sujeira
                for col_limpa in range(10, 22):
                    target_limpa = ws.cell(row=curr_row, column=col_limpa)
                    target_limpa.value = None
                    target_limpa.fill = openpyxl.styles.PatternFill(fill_type=None)

            linhas_modificadas += 1

print(f"\nFinalizado! {linhas_modificadas} áreas de Headbutt foram divididas.")
print("Salvando o novo arquivo... Aguarde.")

# Salva como um arquivo novo para você testar se ficou perfeito
wb.save("SoulSilverOficial_Separado.xlsm")
print("Salvo com sucesso no arquivo 'SoulSilverOficial_Separado.xlsm'!")