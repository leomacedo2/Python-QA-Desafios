import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import time

pokemon_cache = {}
species_varieties_cache = {}

# --- SUAS REGRAS PESSOAIS ---
NOTAS_PESSOAIS = {
    "Gloom": "Vileplume: Macho. Bellossom: Fêmea",
    "Poliwhirl": "Poliwrath: Macho. Politoed: Fêmea",
    "Slowpoke": "Slowking: Macho. Slowbro: Fêmea",
    "Tyrogue": "Pedras Evolutivas pros 3",
    "Eevee": "Pedras Evolutivas pros 7",
    "Pichu": "Tenta ter os dois!",
    "Unown": "Tenta ter todos mas só use 1!",
    "Wurmple": "Cascoon: Macho. Silcoon: Fêmea",
    "Kirlia": "Gallade: Macho. Gardevoir: Fêmea",
    "Nincada": "Ninjask: Macho. Shedinja: Fêmea",
    "Castform": "Tenta ter um com os golpes de mudança de clima",
    "Snorunt": "Glalie: Macho. Froslass: Fêmea",
    "Clamperl": "Huntail: Macho. Gorebyss: Fêmea",
    "Deoxys": "Tenha as 4 formas!",
    "Burmy": "Mantenha o padrão / O padrão muda conforme o local!",
    "Wormadam": "Tente ter as 3 formas adquiridas de evolução!",
    "Cherrim": "Só um Cherrim com o golpe Sunny Day!",
    "Shellos": "Tentar ter os dois!",
    "Gastrodon": "Tentar ter os dois!",
    "Rotom": "Ter as 6 formas! Adquira 6 rotoms e evolua quando achar os trecos",
    "Giratina": "Tenha as 2 formas!",
    "Shaymin": "Aprender os exclusivos das duas em um só! Muda dia/noite",
    "Arceus": "Só precisa de um Arceus e as 18 placas!"
}

GOLPES_DORMIR = {"Hypnosis", "Sleep Powder", "Spore", "Sing", "Yawn", "Lovely Kiss", "Grass Whistle", "Dark Void"}

def get_pokemon_data(identifier):
    if identifier in pokemon_cache:
        return pokemon_cache[identifier]
        
    url = f"https://pokeapi.co/api/v2/pokemon/{identifier}"
    try:
        res = requests.get(url)
        if res.status_code != 200: return None
        data = res.json()
    except:
        return None

    base_name = data['species']['name'].title().replace('-', ' ')
    name = data['name'].title().replace('-', ' ')
    
    moves_list = []
    lvl1_moves = set()
    moves_by_lvl = {}
    
    for item in data['moves']:
        move_name = item['move']['name'].title().replace('-', ' ')
        for v in item['version_group_details']:
            if v['version_group']['name'] == 'heartgold-soulsilver' and v['move_learn_method']['name'] == 'level-up':
                lvl = v['level_learned_at']
                if lvl > 0:
                    moves_list.append((lvl, move_name))
                    
                    if lvl == 1:
                        lvl1_moves.add(move_name)
                    
                    if lvl > 1:
                        if lvl not in moves_by_lvl:
                            moves_by_lvl[lvl] = []
                        if move_name not in moves_by_lvl[lvl]:
                            moves_by_lvl[lvl].append(move_name)
                            
    if not moves_list:
        pokemon_cache[identifier] = None
        return None
                    
    moves_list.sort(key=lambda x: x[0])
    
    simultaneous_moves = {lvl: moves for lvl, moves in moves_by_lvl.items() if len(moves) > 1}
    all_move_names = set([m[1] for m in moves_list])
    later_moves = set([m[1] for m in moves_list if m[0] > 1])
    repeated_lvl1 = {m for m in lvl1_moves if m in later_moves}
    
    pokemon_cache[identifier] = {
        'id': data['id'], 
        'name': name,
        'base_name': base_name,
        'moves': moves_list,
        'all_move_names': all_move_names,
        'lvl1_moves': lvl1_moves,      # NOVO: Guardando o set puro de Lvl 1
        'later_moves': later_moves,    # NOVO: Guardando o set puro de Lvl > 1
        'lvl1_count': len(lvl1_moves),
        'simultaneous': simultaneous_moves,
        'repeated_lvl1': repeated_lvl1
    }
    return pokemon_cache[identifier]

def get_hgss_varieties(species_id):
    if species_id in species_varieties_cache: return species_varieties_cache[species_id]
    
    res = requests.get(f"https://pokeapi.co/api/v2/pokemon-species/{species_id}/")
    if res.status_code != 200: return []
    data = res.json()
    
    varieties = []
    default_moves = None
    
    for v in data['varieties']:
        if v['is_default']:
            default_var_name = v['pokemon']['name']
            p_data = get_pokemon_data(default_var_name)
            if p_data and len(p_data['moves']) > 0:
                default_moves = tuple(p_data['moves'])
                varieties.append(default_var_name)
            break
            
    for v in data['varieties']:
        if v['is_default']: continue
        var_name = v['pokemon']['name']
        p_data = get_pokemon_data(var_name)
        
        if p_data and len(p_data['moves']) > 0:
            if default_moves and tuple(p_data['moves']) == default_moves:
                continue 
            varieties.append(var_name)
            
    species_varieties_cache[species_id] = varieties
    return varieties

def traduzir_metodo_evolucao(detalhes):
    if not detalhes: return "Base", "base"
    d = detalhes[0]
    trigger = d['trigger']['name']
    
    is_especial = False
    gender_str = ""
    
    if d.get('gender') == 1:
        gender_str = " (Apenas Fêmea)"
        is_especial = True
    elif d.get('gender') == 2:
        gender_str = " (Apenas Macho)"
        is_especial = True

    if trigger == 'level-up':
        condicoes = []
        if d.get('min_level'): condicoes.append(f"Lvl {d['min_level']}")
        else:
            condicoes.append("Lvl Up")
            is_especial = True 
            
        if d.get('min_happiness'): condicoes.append("Felicidade Alta"); is_especial = True
        if d.get('min_beauty'): condicoes.append("Beleza Alta"); is_especial = True
        if d.get('min_affection'): condicoes.append("Afeição Alta"); is_especial = True
        if d.get('time_of_day'):
            tempo = d['time_of_day']
            if tempo == 'day': condicoes.append("de Dia")
            elif tempo == 'night': condicoes.append("de Noite")
            elif tempo == 'dusk': condicoes.append("no Crepúsculo")
            is_especial = True
        if d.get('held_item'): condicoes.append(f"segurando {d['held_item']['name'].title().replace('-', ' ')}"); is_especial = True
        if d.get('known_move'): condicoes.append(f"sabendo o golpe {d['known_move']['name'].title().replace('-', ' ')}"); is_especial = True
        if d.get('known_move_type'): condicoes.append(f"sabendo golpe do tipo {d['known_move_type']['name'].title()}"); is_especial = True
        if d.get('location'): condicoes.append(f"em {d['location']['name'].title().replace('-', ' ')}"); is_especial = True
        if d.get('needs_overworld_rain'): condicoes.append("chovendo no mapa"); is_especial = True
        if d.get('party_species'): condicoes.append(f"com {d['party_species']['name'].title()} na Party"); is_especial = True
        if d.get('party_type'): condicoes.append(f"com Pokémon tipo {d['party_type']['name'].title()} na Party"); is_especial = True
        if d.get('relative_physical_stats') is not None:
            rps = d['relative_physical_stats']
            if rps == 1: condicoes.append("(Atq > Def)")
            elif rps == -1: condicoes.append("(Atq < Def)")
            elif rps == 0: condicoes.append("(Atq = Def)")
            is_especial = True
        if d.get('turn_upside_down'): condicoes.append("(Console de cabeça pra baixo!)"); is_especial = True
            
        texto_final = " + ".join(condicoes) + gender_str
        tipo_trigger = "especial" if is_especial else "normal"
        return texto_final, tipo_trigger
        
    elif trigger == 'use-item':
        item_name = d['item']['name'].title().replace('-', ' ')
        return f"Usar item: {item_name}{gender_str}", "especial"
        
    elif trigger == 'trade':
        trade_str = "Troca"
        if d.get('held_item'): trade_str += f" segurando {d['held_item']['name'].title().replace('-', ' ')}"
        if d.get('trade_species'): trade_str += f" por {d['trade_species']['name'].title()}"
        return trade_str + gender_str, "especial"
        
    elif trigger == 'shed': return "Lvl 20 + Espaço livre + Pokébola", "especial"
    elif trigger == 'spin': return "Girar o personagem", "especial"

    return f"{trigger.title()}{gender_str}", "especial"

def get_evolution_paths(chain_node, evo_method="Base", evo_trigger="base"):
    species_url = chain_node['species']['url']
    species_id = int(species_url.strip('/').split('/')[-1])
    
    varieties = get_hgss_varieties(species_id)
    if not varieties: return []
    
    nodes = []
    for var_name in varieties:
        nodes.append({'identifier': var_name, 'evo_method': evo_method, 'evo_trigger': evo_trigger})
        
    if not chain_node['evolves_to']: return [[n] for n in nodes]
    
    paths = []
    for evo in chain_node['evolves_to']:
        next_evo_method, next_trigger = traduzir_metodo_evolucao(evo['evolution_details'])
        sub_paths = get_evolution_paths(evo, next_evo_method, next_trigger)
        for p in sub_paths:
            for n in nodes:
                paths.append([n] + p)
    return paths

def checar_combos(all_moves):
    alertas = []
    if "Stockpile" in all_moves:
        tem_spit = "Aprende" if "Spit Up" in all_moves else "NÃO aprende"
        tem_swallow = "Aprende" if "Swallow" in all_moves else "NÃO aprende"
        alertas.append(f"Usa Stockpile: ({tem_spit} Spit Up / {tem_swallow} Swallow)")
        
    if "Snore" in all_moves or "Sleep Talk" in all_moves:
        tem_rest = "Aprende" if "Rest" in all_moves else "NÃO aprende"
        alertas.append(f"Usa Snore/Sleep Talk: ({tem_rest} Rest no Level Up)")
        
    if "Dream Eater" in all_moves or "Nightmare" in all_moves:
        golpes_aprendidos = [m for m in GOLPES_DORMIR if m in all_moves]
        if golpes_aprendidos:
            nomes_golpes = ", ".join(golpes_aprendidos)
            alertas.append(f"Usa Dream Eater/Nightmare: (Com: {nomes_golpes})")
        else:
            alertas.append("Usa Dream Eater/Nightmare: (SEM golpe de dormir no Level Up)")
        
    if not alertas: return "-"
    return " | ".join(alertas)

def gerar_planilha_horizontal():
    print("Iniciando a montagem da Planilha Definitiva...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Familias HGSS"
    
    # --- CORES ---
    fill_nao_universal = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid") 
    fill_nao_universal_forte = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid") # NOVO: Amarelo forte para 2+ faltantes!
    fill_simultaneo = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid") 
    fill_repetido = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid") 
    fill_relearn_exclusivo = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid") # NOVO: Ciano para Relearn!
    
    fill_verde_bloco = PatternFill(start_color="C3E6CB", end_color="C3E6CB", fill_type="solid") 
    fill_rosa_bloco = PatternFill(start_color="FF99CC", end_color="FF99CC", fill_type="solid") 
    fill_vermelho_forte = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") 
    fill_roxo_claro = PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")
    fill_laranja_claro = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
    fill_azul_forte = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    
    font_bold = Font(bold=True)
    font_branca_bold = Font(bold=True, color="FFFFFF")
    font_branca = Font(color="FFFFFF")
    
    headers = ['ID', 'Pokemon', 'Level', 'Golpe'] * 3
    ws.append(headers)
    for col_num in range(1, 13):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font_bold
        cell.alignment = Alignment(horizontal='center')

    current_row = 2
    
    for chain_id in range(1, 260):
        url = f"https://pokeapi.co/api/v2/evolution-chain/{chain_id}/"
        try:
            res = requests.get(url)
            if res.status_code != 200: continue
            chain_data = res.json()
        except: continue
            
        paths = get_evolution_paths(chain_data['chain'])
        
        for path_nodes in paths:
            path_data = []
            for node in path_nodes:
                data = get_pokemon_data(node['identifier'])
                if data:
                    pkmn_info = data.copy()
                    pkmn_info['evo_method'] = node['evo_method']
                    pkmn_info['evo_trigger'] = node['evo_trigger']
                    path_data.append(pkmn_info)
            
            if not path_data: continue
            print(f"Processando família: {' -> '.join([p['name'] for p in path_data])}")
            
            max_moves_len = max([len(p['moves']) for p in path_data] + [1])
            list_of_move_sets = [p['all_move_names'] for p in path_data]
            universal_moves = set.intersection(*list_of_move_sets) if list_of_move_sets else set()
            
            base_pkmn = path_data[0] # Pokémon Base da Família
            max_info_rows_needed = 1
            
            for stage_idx, pkmn in enumerate(path_data):
                col_base = (stage_idx * 4) + 1 
                
                # --- MATEMÁTICA DO RELEARN EXCLUSIVO ---
                relearn_exclusives = set()
                if pkmn['lvl1_count'] >= 5:
                    relearn_exclusives = (pkmn['lvl1_moves'] - pkmn['later_moves']) - base_pkmn['all_move_names']
                
                # Preenchendo golpes
                for r_offset in range(max_moves_len):
                    row_idx = current_row + r_offset
                    ws.cell(row=row_idx, column=col_base, value=pkmn['id']).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_idx, column=col_base + 1, value=pkmn['name'])
                    
                    if r_offset < len(pkmn['moves']):
                        lvl, move_name = pkmn['moves'][r_offset]
                        display_name = move_name
                        nao_aprendem = []
                        
                        # Pegamos quem não aprende para usar tanto no texto quanto na cor
                        if move_name not in universal_moves:
                            nao_aprendem = [p['name'] for p in path_data if move_name not in p['all_move_names']]
                            
                            # Adicionando texto curto para famílias de 3+ membros
                            if len(path_data) >= 3 and nao_aprendem:
                                display_name = f"{move_name} ({', '.join(nao_aprendem)})"

                        cell_lvl = ws.cell(row=row_idx, column=col_base + 2, value=lvl)
                        cell_lvl.alignment = Alignment(horizontal='right')
                        cell_move = ws.cell(row=row_idx, column=col_base + 3, value=display_name)
                        
                        # --- APLICANDO AS CORES ---
                        if lvl in pkmn['simultaneous']:
                            cell_lvl.fill = fill_simultaneo; cell_move.fill = fill_simultaneo
                            
                        # NOVA LÓGICA: Amarelo Claro (1 forma falta) vs Amarelo Forte (2+ faltam)
                        if move_name not in universal_moves:
                            if len(nao_aprendem) >= 2:
                                cell_move.fill = fill_nao_universal_forte
                            else:
                                cell_move.fill = fill_nao_universal
                                
                        if (lvl == 1) and (move_name in pkmn['repeated_lvl1']):
                            cell_move.fill = fill_repetido
                            
                        # Ciano preenchendo a linha do LEVEL
                        if (lvl == 1) and (move_name in relearn_exclusives):
                            cell_lvl.fill = fill_relearn_exclusivo

                # --- BLOCO DE ANÁLISE ---
                info_row_iter = current_row + max_moves_len
                
                evo_fill = fill_verde_bloco if pkmn['evo_trigger'] in ['normal', 'base'] else fill_rosa_bloco
                c_evo_lbl = ws.cell(row=info_row_iter, column=col_base + 2, value="Como Evoluir:")
                c_evo_val = ws.cell(row=info_row_iter, column=col_base + 3, value=pkmn['evo_method'])
                c_evo_lbl.font = font_bold; c_evo_lbl.fill = evo_fill; c_evo_val.fill = evo_fill
                info_row_iter += 1
                
                if pkmn['lvl1_count'] > 4:
                    c_lvl1_lbl = ws.cell(row=info_row_iter, column=col_base + 2, value="Relearn?")
                    
                    if relearn_exclusives:
                        texto_relearn = f"Sim! Especiais: {', '.join(sorted(list(relearn_exclusives)))}"
                        c_lvl1_lbl.fill = fill_azul_forte # Azul escuro pra dar destaque nos Especiais
                        c_lvl1_val = ws.cell(row=info_row_iter, column=col_base + 3, value=texto_relearn)
                        c_lvl1_val.fill = fill_azul_forte
                    else:
                        texto_relearn = f"Sim ({pkmn['lvl1_count']} golpes padrão)"
                        c_lvl1_lbl.fill = fill_vermelho_forte
                        c_lvl1_val = ws.cell(row=info_row_iter, column=col_base + 3, value=texto_relearn)
                        c_lvl1_val.fill = fill_vermelho_forte
                        
                    c_lvl1_lbl.font = font_branca_bold; c_lvl1_val.font = font_branca
                    info_row_iter += 1

                nota = NOTAS_PESSOAIS.get(pkmn['base_name'])
                if not nota: nota = NOTAS_PESSOAIS.get(pkmn['name']) 
                    
                if nota:
                    c_nota_lbl = ws.cell(row=info_row_iter, column=col_base + 2, value="Sua Regra:")
                    c_nota_val = ws.cell(row=info_row_iter, column=col_base + 3, value=nota)
                    c_nota_lbl.fill = fill_roxo_claro; c_nota_lbl.font = font_bold; c_nota_val.fill = fill_roxo_claro
                    info_row_iter += 1
                
                alerta = checar_combos(pkmn['all_move_names'])
                if alerta != "-":
                    c_alerta_lbl = ws.cell(row=info_row_iter, column=col_base + 2, value="Alerta Combo:")
                    c_alerta_val = ws.cell(row=info_row_iter, column=col_base + 3, value=alerta)
                    c_alerta_lbl.fill = fill_laranja_claro; c_alerta_lbl.font = font_bold; c_alerta_val.fill = fill_laranja_claro
                    info_row_iter += 1

                rows_used = info_row_iter - (current_row + max_moves_len)
                if rows_used > max_info_rows_needed:
                    max_info_rows_needed = rows_used

            current_row += max_moves_len + max_info_rows_needed + 2
            time.sleep(0.1)

    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(cell.value)
            except: pass
        ws.column_dimensions[col[0].column_letter].width = (max_length + 2)

    nome_arquivo = 'Golpes_Familias_Definitiva_HGSS.xlsx'
    wb.save(nome_arquivo)
    print(f"\n🎉 SUCESSO! Arquivo salvo com os Relearns Exclusivos identificados!")

if __name__ == "__main__":
    gerar_planilha_horizontal()